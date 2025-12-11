from fastapi import APIRouter, HTTPException, UploadFile, File, status
from typing import List
from app.crud.crud_user import user as crud_user
from app.schemas.user_schema import UserCreate, UserUpdate, UserResponse
from app.models.user_model import UserRole
from app.core.database import get_database
from app.core.security import get_password_hash
import openpyxl
from io import BytesIO
from bson import ObjectId

router = APIRouter()

@router.post("/import-padres", status_code=status.HTTP_201_CREATED)
async def import_padres(file: UploadFile = File(...)):
    """
    Importar padres masivamente desde Excel.
    Columnas: Email, Password, Nombre, Apellido, Telefono
    """
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="El archivo debe ser un Excel (.xlsx)")

    contents = await file.read()
    workbook = openpyxl.load_workbook(BytesIO(contents))
    sheet = workbook.active

    db = get_database()
    creados = []
    errores = []

    for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        try:
            # Asumimos orden: Email, Password, Nombre, Apellido, Telefono
            if not row[0] or not row[1]: 
                continue

            email = row[0]
            password = row[1]
            nombre = row[2]
            apellido = row[3]
            telefono = row[4]

            # Verificar si existe
            existing = await crud_user.get_by_email(db, email=email)
            if existing:
                errores.append(f"Fila {index}: Email {email} ya existe")
                continue

            user_in = UserCreate(
                email=email,
                password=get_password_hash(str(password)),
                nombre=str(nombre) if nombre else None,
                apellido=str(apellido) if apellido else None,
                telefono=str(telefono) if telefono else None,
                role=UserRole.PADRE,
                hijos_ids=[] # Sin vincular hijos automáticamente
            )
            
            nuevo_user = await crud_user.create(db, obj_in=user_in)
            creados.append(nuevo_user)

        except Exception as e:
            errores.append(f"Fila {index}: Error - {str(e)}")

    return {
        "message": "Importación de padres finalizada",
        "creados_count": len(creados),
        "errores": errores
    }

@router.post("/bulk-delete-padres", status_code=status.HTTP_200_OK)
async def bulk_delete_padres(file: UploadFile = File(...)):
    """
    Eliminar padres masivamente basado en Email del Excel.
    """
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="El archivo debe ser un Excel (.xlsx)")

    contents = await file.read()
    workbook = openpyxl.load_workbook(BytesIO(contents))
    sheet = workbook.active

    db = get_database()
    collection = db["users"]
    eliminados = 0
    errores = []

    for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        try:
            if not row[0]: # Email
                continue
            
            email = row[0]
            
            # Borrar por Email
            result = await collection.delete_one({"email": email})
            if result.deleted_count > 0:
                eliminados += 1
            else:
                errores.append(f"Fila {index}: Email {email} no encontrado")

        except Exception as e:
            errores.append(f"Fila {index}: Error - {str(e)}")

    return {
        "message": "Eliminación masiva de padres finalizada",
        "eliminados_count": eliminados,
        "errores": errores
    }

# --- Standard CRUD ---

@router.get("/", response_model=List[UserResponse])
async def read_users(skip: int = 0, limit: int = 100):
    db = get_database()
    return await crud_user.get_multi(db, skip=skip, limit=limit)

@router.post("/", response_model=UserResponse)
async def create_user(user_in: UserCreate):
    db = get_database()
    # Hash password if provided
    if user_in.password:
        user_in.password = get_password_hash(user_in.password)
    return await crud_user.create(db, obj_in=user_in)

@router.get("/{id}", response_model=UserResponse)
async def read_user(id: str):
    db = get_database()
    user = await crud_user.get(db, id=id)
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    return user

@router.put("/{id}", response_model=UserResponse)
async def update_user(id: str, user_in: UserUpdate):
    db = get_database()
    user = await crud_user.get(db, id=id)
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    if user_in.password:
        user_in.password = get_password_hash(user_in.password)
    return await crud_user.update(db, db_obj=user, obj_in=user_in)

@router.delete("/{id}", response_model=UserResponse)
async def delete_user(id: str):
    db = get_database()
    user = await crud_user.get(db, id=id)
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    return await crud_user.remove(db, id=id)
