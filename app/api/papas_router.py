from fastapi import APIRouter, HTTPException, UploadFile, File, status, Query, Body
from typing import List, Optional
import math
from app.crud.crud_papa import papa as crud_papa
from app.schemas.papa_schema import PapaCreate, PapaUpdate, PapaResponse
from app.schemas.common import PaginatedResponse
from app.models.common import UserRole
from app.core.database import get_database
from app.core.security import get_password_hash
import openpyxl
from io import BytesIO
from bson import ObjectId

from app.models.malla_curricular_model import NivelEducativo
from app.models.curso_model import TurnoCurso
from app.schemas.estudiante_schema import GradoFilter

router = APIRouter()

@router.post("/import-padres", status_code=status.HTTP_201_CREATED)
async def import_padres(file: UploadFile = File(...)):
    """
    Importar padres masivamente desde Excel (Router Papas).
    Columnas: Email, Password, Nombre, Apellido, Telefono
    """
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="El archivo debe ser un Excel (.xlsx)")

    contents = await file.read()
    workbook = openpyxl.load_workbook(BytesIO(contents))
    sheet = workbook.active

    db = get_database()
    creados = []
    errores = []

    for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        try:
            # Asumimos orden: Email, Password, Nombre, Apellido, Telefono
            if not row[0] or not row[1]: 
                continue

            email = row[0]
            password = row[1]
            nombre = row[2]
            apellido = row[3]
            telefono = row[4]

            # Verificar si existe
            existing = await crud_papa.get_by_email(db, email=email)
            if existing:
                errores.append(f"Fila {index}: Email {email} ya existe")
                continue

            # Crear objeto PapaCreate
            user_in = PapaCreate(
                email=email,
                password=get_password_hash(str(password)),
                nombre=str(nombre) if nombre else "Sin Nombre",
                apellido=str(apellido) if apellido else "Sin Apellido",
                telefono=str(telefono) if telefono else None,
                role=UserRole.PADRE,
                hijos_ids=[] 
            )
            
            # create maneja el mapeo password -> hashed_password si se pasa plano
            # Aqui pasamos ya hasheado en password field. 
            # OJO: crud_papa.create espera password plano o lo que sea que venga en 'password' field y lo mueve a hashed_password.
            # Si pasamos hash en 'password', se guardará como 'hashed_password'. Correcto.
            
            nuevo_user = await crud_papa.create(db, obj_in=user_in)
            creados.append(nuevo_user)

        except Exception as e:
            errores.append(f"Fila {index}: Error - {str(e)}")

    return {
        "message": "Importación de padres finalizada",
        "creados_count": len(creados),
        "errores": errores
    }

@router.post("/bulk-delete-padres", status_code=status.HTTP_200_OK)
async def bulk_delete_padres(file: UploadFile = File(...)):
    """
    Eliminar padres masivamente basado en Email del Excel.
    """
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="El archivo debe ser un Excel (.xlsx)")

    contents = await file.read()
    workbook = openpyxl.load_workbook(BytesIO(contents))
    sheet = workbook.active

    db = get_database()
    collection = db["users"]
    eliminados = 0
    errores = []

    for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        try:
            if not row[0]: # Email
                continue
            
            email = row[0]
            
            # Borrar por Email
            result = await collection.delete_one({"email": email})
            if result.deleted_count > 0:
                eliminados += 1
            else:
                errores.append(f"Fila {index}: Email {email} no encontrado")

        except Exception as e:
            errores.append(f"Fila {index}: Error - {str(e)}")

    return {
        "message": "Eliminación masiva de padres finalizada",
        "eliminados_count": eliminados,
        "errores": errores
    }

# --- Papas CRUD ---

@router.get("/", response_model=PaginatedResponse[PapaResponse])
async def read_papas_list(
    page: int = Query(1, ge=1),
    per_page: int = Query(10, ge=1, le=100),
    q: Optional[str] = None,
    nivel: Optional[NivelEducativo] = Query(None, description="Filtro por Nivel Educativo"),
    grado: Optional[GradoFilter] = Query(None, description="Filtro por Grado"),
    turno: Optional[TurnoCurso] = Query(None, description="Filtro por Turno"),
    paralelo: Optional[str] = Query(None, description="Filtro por Paralelo (A, B)")
):
    """Listar todos los padres (Role = PADRE) con paginación y búsqueda"""
    db = get_database()
    items, total = await crud_papa.get_paginated(
        db, 
        page=page, 
        per_page=per_page, 
        q=q,
        nivel=nivel,
        grado=grado,
        turno=turno,
        paralelo=paralelo
    )
    
    # Calcular total de páginas
    total_pages = math.ceil(total / per_page) if per_page > 0 else 0
    
    return {
        "total": total,
        "page": page,
        "per_page": per_page,
        "total_pages": total_pages,
        "data": items
    }

@router.post("/", response_model=PapaResponse, status_code=status.HTTP_201_CREATED)
async def create_papa(user_in: PapaCreate):
    """Crear un padre (Rol PADRE forzado)"""
    db = get_database()
    
    # Check email
    existing = await crud_papa.get_by_email(db, email=user_in.email)
    if existing:
        raise HTTPException(status_code=400, detail="El email ya está registrado")

    if user_in.password:
        user_in.password = get_password_hash(user_in.password)
    
    # Force Role
    user_in.role = UserRole.PADRE
    
    return await crud_papa.create(db, obj_in=user_in)

@router.get("/{id}", response_model=PapaResponse)
async def read_papa(id: str):
    """Obtener un padre por ID"""
    db = get_database()
    user = await crud_papa.get(db, id=id)
    # Check Role is PADRE? Yes
    # But user object returned is PapaModel.
    if not user or getattr(user, "role", None) != UserRole.PADRE:
        raise HTTPException(status_code=404, detail="Padre no encontrado")
    return user

@router.put("/{id}", response_model=PapaResponse)
async def update_papa(id: str, user_in: PapaUpdate):
    """Actualizar un padre"""
    db = get_database()
    user = await crud_papa.get(db, id=id)
    if not user or getattr(user, "role", None) != UserRole.PADRE:
        raise HTTPException(status_code=404, detail="Padre no encontrado")
    
    if user_in.password:
        user_in.password = get_password_hash(user_in.password)
        
    return await crud_papa.update(db, db_obj=user, obj_in=user_in)

@router.delete("/{id}", response_model=PapaResponse)
async def delete_papa(id: str):
    """Eliminar un padre"""
    db = get_database()
    user = await crud_papa.get(db, id=id)
    if not user or getattr(user, "role", None) != UserRole.PADRE:
        raise HTTPException(status_code=404, detail="Padre no encontrado")
    
    return await crud_papa.remove(db, id=id)

@router.post("/{id}/hijos", response_model=PapaResponse)
async def assign_child(id: str, child_id: str = Body(..., embed=True)):
    """
    Asignar un hijo a un padre (sin borrar los anteriores).
    Body: {"child_id": "..."}
    """
    db = get_database()
    
    # Verificar que el padre existe
    user = await crud_papa.get(db, id=id)
    if not user:
        raise HTTPException(status_code=404, detail="Padre no encontrado")
        
    updated_papa = await crud_papa.add_child(db, papa_id=id, child_id=child_id)
    return updated_papa

@router.delete("/{id}/hijos/{child_id}", response_model=PapaResponse)
async def unassign_child(id: str, child_id: str):
    """
    Desvincular un hijo de un padre.
    """
    db = get_database()
    
    # Verificar que el padre existe
    user = await crud_papa.get(db, id=id)
    if not user:
        raise HTTPException(status_code=404, detail="Padre no encontrado")
        
    updated_papa = await crud_papa.remove_child(db, papa_id=id, child_id=child_id)
    return updated_papa
