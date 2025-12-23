from typing import List, Optional
from fastapi import APIRouter, HTTPException, UploadFile, File, status, Query, Depends
import math
from app.schemas.common import PaginatedResponse
from app.crud.crud_estudiante import estudiante as crud_estudiante
from app.schemas.estudiante_schema import EstudianteCreate, EstudianteUpdate, EstudianteResponse, GradoFilter, HijoSimpleResponse
from app.models.malla_curricular_model import NivelEducativo
from app.models.curso_model import TurnoCurso
from app.core.database import get_database
import openpyxl
from io import BytesIO
from bson import ObjectId
from app.models.common import UserRole
from app.api.auth_router import get_current_user

router = APIRouter()

@router.post("/import", status_code=status.HTTP_201_CREATED)
async def import_estudiantes(file: UploadFile = File(...)):
    """
    Importar estudiantes masivamente desde Excel.
    Columnas: RUDE, Nombres, Apellidos, Curso ID (Opcional), Estado (Opcional)
    """
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="El archivo debe ser un Excel (.xlsx)")

    contents = await file.read()
    workbook = openpyxl.load_workbook(BytesIO(contents))
    sheet = workbook.active

    db = get_database()
    creados = []
    errores = []

    for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        try:
            # Asumimos orden: RUDE, Nombres, Apellidos, CursoID, Estado
            if not row[0]: 
                continue

            rude = row[0]
            nombres = row[1]
            apellidos = row[2]
            curso_id = row[3] if len(row) > 3 else None
            estado = row[4] if len(row) > 4 else "ACTIVO"

            # Validar RUDE
            # Intentar encontrar si ya existe
            existing = await crud_estudiante.get_multi(db, limit=1) 
            # El CRUD genérico no tiene filtro por campo, deberíamos verificar manualmente en la colección o mejorar el CRUD.
            # Accederemos directo a la colección para chequear duplicidad de RUDE rapidámente.
            collection = db["estudiantes"]
            if await collection.find_one({"rude": rude}):
                errores.append(f"Fila {index}: RUDE {rude} ya existe")
                continue

            # Validar Curso ID
            if curso_id and not ObjectId.is_valid(str(curso_id)):
                errores.append(f"Fila {index}: Curso ID inválido")
                continue

            estudiante_in = EstudianteCreate(
                rude=rude,
                nombres=str(nombres),
                apellidos=str(apellidos),
                curso_id=str(curso_id) if curso_id else None,
                estado=str(estado) if estado else "ACTIVO"
            )

            nuevo_est = await crud_estudiante.create(db, obj_in=estudiante_in)
            creados.append(nuevo_est)

        except Exception as e:
            errores.append(f"Fila {index}: Error - {str(e)}")

    return {
        "message": "Importación finalizada",
        "creados_count": len(creados),
        "errores": errores
    }

@router.post("/bulk-delete", status_code=status.HTTP_200_OK)
async def bulk_delete_estudiantes(file: UploadFile = File(...)):
    """
    Eliminar estudiantes masivamente basado en el RUDE del Excel.
    """
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="El archivo debe ser un Excel (.xlsx)")

    contents = await file.read()
    workbook = openpyxl.load_workbook(BytesIO(contents))
    sheet = workbook.active

    db = get_database()
    collection = db["estudiantes"]
    eliminados = 0
    errores = []

    for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        try:
            if not row[0]: 
                continue
            
            rude = row[0]
            
            # Borrar por RUDE
            result = await collection.delete_one({"rude": rude})
            if result.deleted_count > 0:
                eliminados += 1
            else:
                errores.append(f"Fila {index}: RUDE {rude} no encontrado")

        except Exception as e:
            errores.append(f"Fila {index}: Error - {str(e)}")

    return {
        "message": "Eliminación masiva finalizada",
        "eliminados_count": eliminados,
        "errores": errores
    }

@router.get("/", response_model=PaginatedResponse[EstudianteResponse])
async def read_estudiantes(
    page: int = Query(1, ge=1, description="Número de página"),
    per_page: int = Query(10, ge=1, le=100, description="Registros por página"),
    q: Optional[str] = Query(None, description="Filtro de búsqueda"),
    nivel: Optional[NivelEducativo] = Query(None, description="Filtro por Nivel Educativo"),
    grado: Optional[GradoFilter] = Query(None, description="Filtro por Grado"),
    turno: Optional[TurnoCurso] = Query(None, description="Filtro por Turno"),
    paralelo: Optional[str] = Query(None, description="Filtro por Paralelo (A, B, etc)")
):
    db = get_database()
    items, total = await crud_estudiante.get_multi_paginated(
        db, 
        page=page, 
        per_page=per_page, 
        q=q,
        nivel=nivel,
        grado=grado,
        turno=turno,
        paralelo=paralelo
    )
    
    total_pages = math.ceil(total / per_page) if per_page > 0 else 0
    
    return {
        "total": total,
        "page": page,
        "per_page": per_page,
        "total_pages": total_pages,
        "data": items
    }

@router.post("/", response_model=EstudianteResponse)
async def create_estudiante(estudiante_in: EstudianteCreate):
    db = get_database()
    return await crud_estudiante.create(db, obj_in=estudiante_in)

@router.get("/mis-hijos", response_model=List[HijoSimpleResponse])
async def get_mis_hijos(
    padre_id: Optional[str] = Query(None, description="ID del padre (solo para admins)"),
    current_user: dict = Depends(get_current_user)
):
    """
    Obtener lista de hijos de un padre.
    - Padres: ven solo sus propios hijos (padre_id se ignora)
    - Admins: pueden especificar padre_id para ver hijos de un padre específico, 
              o dejar vacío para ver todos los estudiantes
    Incluye información del curso de cada estudiante.
    """
    db = get_database()
    
    # Si es admin
    if current_user["role"] == UserRole.ADMIN:
        if padre_id:
            # Admin quiere ver hijos de un padre específico
            if not ObjectId.is_valid(padre_id):
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="ID de padre inválido"
                )
            
            # Buscar el padre
            padre = await db["users"].find_one({"_id": ObjectId(padre_id)})
            
            if not padre:
                # Intentar buscar sin importar el tipo de ID
                padre = await db["users"].find_one({"_id": padre_id})
                
            if not padre:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail=f"Padre no encontrado con ID: {padre_id}"
                )
            
            # Verificar que sea un padre
            if padre.get("role") != UserRole.PADRE and padre.get("role") != "PADRE":
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"El usuario especificado no es un padre. Role: {padre.get('role')}"
                )
            
            hijos_ids = padre.get("hijos_ids", [])
            if not hijos_ids:
                return []
            
            # Convertir IDs a ObjectId si son strings
            object_ids = [ObjectId(hid) if isinstance(hid, str) else hid for hid in hijos_ids]
            estudiantes = await db["estudiantes"].find({"_id": {"$in": object_ids}}).to_list(100)
        else:
            # Admin sin padre_id: ver todos los estudiantes
            estudiantes = await db["estudiantes"].find({}).to_list(1000)
    else:
        # Padres solo ven sus propios hijos (ignorar padre_id)
        hijos_ids = current_user.get("hijos_ids", [])
        if not hijos_ids:
            return []
            
        # Convertir IDs a ObjectId si son strings
        object_ids = [ObjectId(hid) if isinstance(hid, str) else hid for hid in hijos_ids]
        
        # Buscar estudiantes
        estudiantes = await db["estudiantes"].find({"_id": {"$in": object_ids}}).to_list(100)
    
    # Simplificar respuesta
    results = []
    for est in estudiantes:
        hijo_data = {
            "_id": est["_id"],
            "nombres": est.get("nombres", ""),
            "apellidos": est.get("apellidos", "")
        }
        
        # Si el estudiante tiene curso_id, buscar la información del curso
        if est.get("curso_id"):
            curso = await db["cursos"].find_one({"_id": ObjectId(est["curso_id"]) if isinstance(est["curso_id"], str) else est["curso_id"]})
            if curso:
                hijo_data["curso"] = {
                    "nombre": curso.get("nombre", ""),
                    "nivel": curso.get("nivel", ""),
                    "turno": curso.get("turno", ""),
                    "paralelo": curso.get("paralelo", "")
                }
        
        results.append(hijo_data)
        
    return results

@router.get("/{id}", response_model=EstudianteResponse)
async def read_estudiante(id: str):
    db = get_database()
    estudiante = await crud_estudiante.get(db, id=id)
    if not estudiante:
        raise HTTPException(status_code=404, detail="Estudiante not found")
    return estudiante

@router.put("/{id}", response_model=EstudianteResponse)
async def update_estudiante(id: str, estudiante_in: EstudianteUpdate):
    db = get_database()
    estudiante = await crud_estudiante.get(db, id=id)
    if not estudiante:
        raise HTTPException(status_code=404, detail="Estudiante not found")
    return await crud_estudiante.update(db, db_obj=estudiante, obj_in=estudiante_in)

@router.delete("/{id}", response_model=EstudianteResponse)
async def delete_estudiante(id: str):
    db = get_database()
    estudiante = await crud_estudiante.get(db, id=id)
    if not estudiante:
        raise HTTPException(status_code=404, detail="Estudiante not found")
    return await crud_estudiante.remove(db, id=id)


