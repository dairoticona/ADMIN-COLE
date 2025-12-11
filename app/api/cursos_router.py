from fastapi import APIRouter, HTTPException, UploadFile, File, status
from typing import List
from app.crud.crud_curso import curso as crud_curso
from app.schemas.curso_schema import CursoCreate, CursoUpdate, CursoResponse
from app.core.database import get_database
import openpyxl
from io import BytesIO
from bson import ObjectId

router = APIRouter()

@router.post("/import", status_code=status.HTTP_201_CREATED)
async def import_cursos(file: UploadFile = File(...)):
    """
    Importar cursos masivamente desde un archivo Excel (.xlsx).
    Columnas requeridas: nombre, paralelo, nivel, turno, malla_id
    """
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="El archivo debe ser un Excel (.xlsx)")

    contents = await file.read()
    workbook = openpyxl.load_workbook(BytesIO(contents))
    sheet = workbook.active

    db = get_database()
    cursos_creados = []
    errores = []

    # Iterar filas (saltando cabecera)
    for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        try:
            # Asumiendo orden: nombre, paralelo, nivel, turno, malla_id, tutor_id(opcional)
            if not row[0]: # Si no hay nombre, saltar
                continue
            
            nombre = row[0]
            paralelo = row[1]
            nivel = row[2]
            turno = row[3]
            malla_id = row[4]
            tutor_id = row[5] if len(row) > 5 else None

            # Validaciones básicas
            if not ObjectId.is_valid(str(malla_id)):
                errores.append(f"Fila {index}: malla_id inválido")
                continue

            curso_in = CursoCreate(
                nombre=str(nombre),
                paralelo=str(paralelo),
                nivel=str(nivel),
                turno=str(turno),
                malla_id=str(malla_id),
                tutor_id=str(tutor_id) if tutor_id and ObjectId.is_valid(str(tutor_id)) else None
            )
            
            nuevo_curso = await crud_curso.create(db, obj_in=curso_in)
            cursos_creados.append(nuevo_curso)
            
        except Exception as e:
            errores.append(f"Fila {index}: Error procesando - {str(e)}")

    return {
        "message": "Proceso de importación finalizado",
        "creados_count": len(cursos_creados),
        "errores": errores
    }

@router.post("/bulk-delete", status_code=status.HTTP_200_OK)
async def bulk_delete_cursos(file: UploadFile = File(...)):
    """
    Eliminar cursos masivamente usando el mismo archivo Excel de importación.
    Busca coincidencias exactas por: nombre, paralelo, nivel, turno, malla_id.
    """
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="El archivo debe ser un Excel (.xlsx)")

    contents = await file.read()
    workbook = openpyxl.load_workbook(BytesIO(contents))
    sheet = workbook.active

    db = get_database()
    collection = db["cursos"]
    eliminados_count = 0
    errores = []

    # Iterar filas (saltando cabecera)
    for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        try:
            if not row[0]: # Si no hay nombre, saltar
                continue
            
            nombre = row[0]
            paralelo = row[1]
            nivel = row[2]
            turno = row[3]
            malla_id = row[4]
            # tutor_id no es crítico para identificar el curso a borrar, lo ignoramos

            # Validaciones básicas de ID
            if not ObjectId.is_valid(str(malla_id)):
                errores.append(f"Fila {index}: malla_id inválido")
                continue

            # Buscar el curso por campos clave
            filter_query = {
                "nombre": str(nombre),
                "paralelo": str(paralelo),
                "nivel": str(nivel),
                "turno": str(turno),
                "malla_id": ObjectId(str(malla_id))
            }
            
            result = await collection.delete_one(filter_query)
            if result.deleted_count > 0:
                eliminados_count += 1
            else:
                errores.append(f"Fila {index}: No se encontró el curso para eliminar")
            
        except Exception as e:
            errores.append(f"Fila {index}: Error procesando - {str(e)}")

    return {
        "message": "Proceso de eliminación masiva finalizado",
        "eliminados_count": eliminados_count,
        "errores": errores
    }

@router.get("/", response_model=List[CursoResponse])
async def read_cursos(skip: int = 0, limit: int = 100):
    db = get_database()
    cursos = await crud_curso.get_multi(db, skip=skip, limit=limit)
    return cursos

@router.post("/", response_model=CursoResponse)
async def create_curso(curso_in: CursoCreate):
    db = get_database()
    return await crud_curso.create(db, obj_in=curso_in)

@router.get("/{id}", response_model=CursoResponse)
async def read_curso(id: str):
    db = get_database()
    curso = await crud_curso.get(db, id=id)
    if not curso:
        raise HTTPException(status_code=404, detail="Curso not found")
    return curso

@router.put("/{id}", response_model=CursoResponse)
async def update_curso(id: str, curso_in: CursoUpdate):
    db = get_database()
    curso = await crud_curso.get(db, id=id)
    if not curso:
        raise HTTPException(status_code=404, detail="Curso not found")
    return await crud_curso.update(db, db_obj=curso, obj_in=curso_in)

@router.delete("/{id}", response_model=CursoResponse)
async def delete_curso(id: str):
    db = get_database()
    curso = await crud_curso.get(db, id=id)
    if not curso:
        raise HTTPException(status_code=404, detail="Curso not found")
    return await crud_curso.remove(db, id=id)
